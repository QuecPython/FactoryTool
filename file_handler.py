#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
@Project:Factory_test 
@File:file_handler.py
@Author:rivern.yuan
@Date:2022/9/7 10:09 
"""
import serial_handler
import openpyxl
import codecs
import json
import sys
import re
import os


class ExcelHandler(object):
    """
    handle test result Excel
    excel_file: filename include full path
    """
    def __init__(self, excel_file):
        self.file = excel_file
        if os.path.exists(self.file):
            pass
        else:
            wb = openpyxl.Workbook()
            wb.save(self.file)
        self.wb = openpyxl.load_workbook(self.file)
        self.ws = self.wb[self.wb.sheetnames[0]]


    def get_rows_columns(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    def get_cell_value(self, row, column):
        return self.ws.cell(row=row, column=column).value

    def get_col_values(self, column):
        rows = self.ws.max_row
        columns_data = []
        for i in range(1, rows + 1):
            columns_data.append(self.ws.cell(row=i, column=column).value)
            return columns_data

    def get_row_values(self, row):
        columns = self.ws.max_column
        rows_data = []
        for i in range(1, columns + 1):
            rows_data.append(self.ws.cell(row=row, column=i).value)
            return rows_data

    def set_cell_value(self, row, column, cell_value):
        try:
            self.ws.cell(row=row, column=column).value = cell_value
        except KeyError as e:
            self.ws.cell(row=row, column=column).value = "Write Fail"
            print(e)
        finally:
            self.wb.save(self.file)

    def close(self):
        self.wb.close()


class JsonHandler(object):
    """
        json_file: filename include full path
    """
    def __init__(self, json_file):
        self._json_name = json_file

    def read_json(self):
        with codecs.open(self._json_name, 'r', 'utf-8') as f:
            data = json.load(f)
        return data

    @staticmethod
    def write_json(data, json_name):
        with codecs.open(json_name, 'w', 'utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


class ScriptHandler(object):
    """
    QuecPython test script
    script_file: filename include full path
    """
    def __init__(self, script_file: str):
        self.script_file = script_file
        self.pattern_py = {re.compile(r'#.*'), re.compile(r'(""".*?""")*|(\'\'\'.*?\'\'\')*'), }

    """compile py script"""
    def script_check(self):
        try:
            if self.script_file[-3:] == ".py":
                with open(self.script_file, "r", encoding='utf-8') as f:
                    script_data = f.read()
                compile(script_data, '', 'exec')
                return True
            else:
                raise TypeError("File %s is not python file!" % self.script_file)
        except:
            compileInfo = sys.exc_info()
            print("测试脚本语法错误" + str(compileInfo[0]) + ":" + str(compileInfo[1]))

    """get effective py cmd"""
    def script_parse(self):
        content = []
        with open(self.script_file, "r") as f:
            for line in f.readlines():
                result = re.sub(line, "", line.split("\n")[0])
                if result != "":
                    content.append(result)
        return content


if __name__ == '__main__':
    excel_test = ExcelHandler("test\\Test-Result.xlsx")
    print(excel_test.get_rows_columns())
    # excel_test.set_cell_value(1, 1, "No.")
    # excel_test.set_cell_value(1, 2, "SerialNum")
    # print(excel_test.get_col_values(1))
    # print(excel_test.get_row_values(1))
    # print(excel_test.get_cell_value(1, 1))

    result = ["com63", "972658346523","128934928734","success", " True ,False"]
    rows, columns = excel_test.get_rows_columns()
    excel_test.set_cell_value(rows + 1, 1, rows)
    for i, value in enumerate(result):
        excel_test.set_cell_value(rows + 1, i + 2, value)


    excel_test.close()

    # py_test = ScriptHandler("module\\module_test.py")
    # if py_test.script_check():
    #     print(py_test.script_parse())
    # ser = serial_handler.SerialHandler("COM63")
    # print(ser.run_cmd(py_test.script_parse()))
